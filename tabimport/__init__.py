"""
tabimport is a Python utility to ease imports of tabular data from CSV, ODF,
XLS or XLSX files.

Usage:
>>> smart_file = FileFactory(file_path)
>>> for data_line in smart_file:
>>>     do_something_with(data_line['header'])

"""
import csv
import logging
import tempfile
from collections import OrderedDict
from datetime import datetime, time

try:
    # XLS format
    # http://www.lexicon.net/sjmachin/xlrd.html
    import xlrd
    has_xlrd = True
except ImportError:
    has_xlrd = False

try:
    # XLSX format
    import openpyxl
    has_openpyxl = True
except ImportError:
    has_openpyxl = False

try:
    # ODF format
    import ooolib
    has_ooolib = True
except ImportError:
    has_ooolib = False

from django.utils.translation import gettext as _

class UnsupportedFileFormat(Exception):
    pass

class HeaderError(Exception):
    pass


class FileFactory:
    """ Returns a file object depending on the file format """
    def __new__(cls, datafile, **imp_params):
        format = cls._sniff_format(datafile)
        if format == 'ods':
            return ODSImportedFile(datafile, **imp_params)
        elif format == 'xls':
            return XLSImportedFile(datafile, **imp_params)
        elif format == 'xlsx':
            return XLSXImportedFile(datafile, **imp_params)
        elif format == 'csv':
            return CSVImportedFile(datafile, **imp_params)
        else:
            raise UnsupportedFileFormat(_("Unknown file extension '%s'") % format)

    @classmethod
    def _sniff_format(cls, dfile):
        """ dfile may be a file path or a django (Uploaded)File object """
        if isinstance(dfile, str):
            format = dfile.rsplit('.', 1)[-1]
        else:
            if "opendocument.spreadsheet" in dfile.content_type or dfile.name.endswith(".ods"):
                format = "ods"
            elif dfile.name.endswith(".xlsx"):
                format = "xlsx"
            elif "excel" in dfile.content_type or dfile.name.endswith(".xls"):
                format = "xls"
            elif "text/csv" in dfile.content_type or dfile.name.endswith(".csv"):
                format = "csv"
            else:
                raise UnsupportedFileFormat(_("This file is not in an accepted format (ods, xls, csv)"))
        return format


class ImportedFile:
    """ Abstract class to get file object in different formats """
    # Set to True if the external lib does not support file-like objects
    force_file_to_disk = False

    def __init__(self, datafile, sheet_index=0, skip_lines=None):
        """ datafile can be either a path (string) or a file object
            sheet_index is the spreadsheet index, if applicable
            skip_lines is a list of row indexes to skip """
        # Some internal variables are initialized in activate_sheet so as they
        # are resetted whenever a new sheet is activated
        self._headers = {} # dict of lists
        self._ignored_headers_idx = {} # dict of lists
        self.data_sheet_indexes = [sheet_index]
        self.skip_lines = skip_lines
        self.file_content = None
        if isinstance(datafile, str):
            self.file_path = datafile
        else:
            try:
                self.file_path = datafile.temporary_file_path()
            except AttributeError:
                try:
                    self.file_path = datafile.file.name
                except AttributeError:
                    if self.force_file_to_disk:
                        # set to self, so it is not removed before the instance is deleted
                        self.temp_f = tempfile.NamedTemporaryFile()
                        for data in datafile.chunks():
                            self.temp_f.write(data)
                        assert self.temp_f.tell() == datafile.size
                        self.file_path = self.temp_f.name
                    else:
                        self.file_path = None
                        self.file_content = datafile.file.read()

    def __iter__(self):
        return self

    def __next__(self):
        raise NotImplementedError("Abstract class")

    def next(self):
        # Python 2 compatibility
        return type(self).__next__(self)

    def get_headers(self):
        raise NotImplementedError("Abstract class")

    def activate_sheet(self, idx):
        self.current_index = idx
        self._row_index = 1 # skip first line

    def current_sheet_name(self):
        raise NotImplementedError("Abstract class")

    def check_header_validity(self, possible_headers, mandatory_headers, case_sensitive=False): # TODO: ignore_pattern
        """ This method has the side effect of swallowing the first line (headers) of the file """
        errors = []
        warnings = [] # List of tuples: (sheet name, header name, motive)
        def idx_to_header(idx):
            return (idx>25 and chr(64 + idx/26) or '') + chr(65 + idx%26)
        for sheet_idx in self.data_sheet_indexes:
            self.activate_sheet(sheet_idx)
            good_headers = set()
            if not case_sensitive:
                possible_headers = [h.lower() for h in possible_headers]
            for i, h in enumerate(self.get_headers()):
                h_norm = case_sensitive and h.strip() or h.strip().lower()
                if h_norm == "":
                    warnings.append(
                        (self.current_sheet_name(),
                         "(%s header empty)" % idx_to_header(i),
                         _("Empty header")))
                    self._ignored_headers_idx[self.current_index].append(i)
                elif h_norm not in possible_headers:
                    warnings.append((self.current_sheet_name(), h, _("Unknown header")))
                    self._ignored_headers_idx[self.current_index].append(i)
                elif h in good_headers:
                    errors.append(_("The column '%s' is twice in your file's headers (sheet '%s')") % (h, self.current_sheet_name()))
                else:
                    good_headers.add(h)
            for h in mandatory_headers:
                if h not in good_headers:
                    errors.append(_("The header '%s' is mandatory and is missing in sheet '%s' of your file") % (h, self.current_sheet_name()))
        self.activate_sheet(0)
        if errors:
            raise HeaderError("\n".join(errors))
        return warnings


class CSVImportedFile(ImportedFile):
    encoding = 'latin-1'
    def __init__(self, datafile, sheet_index=0, skip_lines=None, **kwds):
        super().__init__(datafile, sheet_index)
        if isinstance(datafile, str):
            # if datafile is a path, try to open the file
            datafile = open(datafile, 'r')
        try:
            dialect = csv.Sniffer().sniff(datafile.read(2048))
            # Python csv module weakness?
            if not dialect.delimiter or dialect.delimiter == '\r':
                dialect.delimiter = ";"
        except Exception:
            dialect = csv.excel
            dialect.delimiter = kwds.get('delimiter', ';')
        self.delimiter = dialect.delimiter
        datafile.seek(0)
        self.reader = csv.DictReader(datafile, dialect=dialect, **kwds)
        # It may happen that fieldnames are not filled before first line has been read
        self._first_line_read = False
        self._first_line = None
        self.current_index = 0

    def get_headers(self):
        if not self.current_index in self._headers:
            self._ignored_headers_idx[self.current_index] = []
            if not self._first_line_read:
                self._first_line = next(self.reader)
                self._first_line_read = True
            self._headers[self.current_index] = self.reader.fieldnames
            if not isinstance(self._headers[self.current_index], list):
                self._headers[self.current_index] = self._headers[self.current_index].split(self.delimiter)
        return self._headers[self.current_index]

    def __next__(self):
        """ Returns an OrderedDict : {'DESCRIPTOR': value, ...} """
        if self._first_line:
            row = self._first_line
            self._first_line = None
        else:
            row = next(self.reader)
        for key, val in row.items():
            row[key] = '' if val is None else val
        return row

    def current_sheet_name(self):
        return "1"


class XLSImportedFile(ImportedFile):
    """ XLS reader based on xlrd (multiple sheets not yet implemented)"""
    def __init__(self, datafile, sheet_index=0, skip_lines=None):
        if not has_xlrd:
            raise NotImplementedError("The xlrd library is not available")
        super().__init__(datafile, sheet_index, skip_lines)
        try:
            self.book = xlrd.open_workbook(filename=self.file_path, file_contents=self.file_content)
        except xlrd.XLRDError as e:
            logging.warn("XLS import error: %s" % str(e))
            raise UnsupportedFileFormat(_("Unable to read the file. Are you sure it is an XLS file?"))
        self.data_sheet_indexes = [i for i, ws in enumerate(self.book.sheets()) if (ws.nrows > 0 and ws.ncols > 0)]
        self.activate_sheet(self.data_sheet_indexes[0])

    def get_headers(self):
        if not self.current_index in self._headers:
            self._headers[self.current_index] = []
            self._ignored_headers_idx[self.current_index] = []
            row = self.current_sheet.row(0)
            for i, cell in enumerate(row):
                self._headers[self.current_index].append(str(cell.value).strip())
        return self._headers[self.current_index]

    def __next__(self):
        """ Returns an OrderedDict : {'DESCRIPTOR': value, ...} """
        while self.skip_lines and self._row_index in self.skip_lines:
            self._row_index += 1
        if self._row_index >= self._nrows:
            # Increment current_index and skip to next sheet, if any
            try:
                new_index = self.data_sheet_indexes[self.data_sheet_indexes.index(self.current_index)+1]
            except IndexError:
                pass
            else:
                self.activate_sheet(new_index)
                return next(self)
            raise StopIteration
        row_dict = OrderedDict()
        row = self.current_sheet.row(self._row_index)
        headers = self.get_headers()
        for i, cell in enumerate(row):
            if i in self._ignored_headers_idx[self.current_index] or i >= len(headers):
                continue
            if cell.ctype == xlrd.XL_CELL_DATE:
                date_tuple = xlrd.xldate_as_tuple(cell.value, self.book.datemode)
                if date_tuple[0] == 0:
                    # No year, probably a time value
                    value = time(*date_tuple[3:])
                else:
                    value = datetime(*date_tuple)
            else:
                value = cell.value
            row_dict[headers[i]] = value
        self._row_index += 1
        return row_dict

    def activate_sheet(self, idx):
        super().activate_sheet(idx)
        self.current_sheet = self.book.sheet_by_index(idx)
        self._nrows = self.current_sheet.nrows
        self._ncols = self.current_sheet.ncols

    def current_sheet_name(self):
        return self.book.sheet_by_index(self.current_index).name


class XLSXImportedFile(ImportedFile):
    """ XLS reader based on xlrd (multiple sheets not yet implemented)"""
    def __init__(self, datafile, sheet_index=0, skip_lines=()):
        if not has_openpyxl:
            raise NotImplementedError("The openpyxl library is not available")
        super().__init__(datafile, sheet_index=sheet_index, skip_lines=skip_lines)
        try:
            self.book = openpyxl.load_workbook(filename=self.file_path)
        except Exception as e:
            logging.warn("XLSX import error: %s" % str(e))
            raise UnsupportedFileFormat(_("Unable to read the file. Are you sure it is an XLSX file?"))
        self.data_sheet_indexes = [
            idx for idx, ws in enumerate(self.book.worksheets)
            if (ws.max_row > 1 and ws.max_column > 1)
        ]
        self.activate_sheet(self.data_sheet_indexes[0])

    def get_headers(self):
        if not self.current_index in self._headers:
            self._headers[self.current_index] = []
            self._ignored_headers_idx[self.current_index] = []
            row = next(self.current_sheet.rows)
            for i, cell in enumerate(row):
                self._headers[self.current_index].append(str(cell.value).strip())
        return self._headers[self.current_index]

    def __next__(self):
        """ Returns an OrderedDict : {'DESCRIPTOR': value, ...} """
        while self.skip_lines and self._row_index in self.skip_lines:
            self._row_index += 1
        try:
            row = next(self._row_iterator)
        except StopIteration:
            # Increment current_index and skip to next sheet, if any
            try:
                new_index = self.data_sheet_indexes[self.data_sheet_indexes.index(self.current_index)+1]
            except (IndexError, ValueError):
                pass
            else:
                self.activate_sheet(new_index)
                return next(self)
            raise
        row_dict = OrderedDict()
        headers = self.get_headers()
        for i, cell in enumerate(row):
            if i in self._ignored_headers_idx[self.current_index] or i >= len(headers):
                continue
            row_dict[headers[i]] = cell.value
        self._row_index += 1
        return row_dict

    def activate_sheet(self, idx):
        super().activate_sheet(idx)
        self.current_sheet = self.book.worksheets[idx]
        self._row_iterator = self.current_sheet.iter_rows(min_row=2)

    def current_sheet_name(self):
        return self.book.sheetnames[self.current_index]


class ODSImportedFile(ImportedFile):
    """ OO Calc reader based on ooolib (multiple sheets not yet implemented)"""
    # FIXME: missing multi-sheet import
    def __init__(self, datafile, sheet_index=0, skip_lines=None):
        if not has_ooolib:
            raise NotImplementedError("The ooolib library is not available")
        super().__init__(datafile, sheet_index, skip_lines)
        book = ooolib.Calc(opendoc=self.file_path)
        book.set_sheet_index(sheet_index)
        self.current_sheet = book
        (self._ncols, self._nrows) = self.current_sheet.get_sheet_dimensions()

    def get_headers(self):
        if not self.current_index in self._headers:
            self._headers[self.current_index] = []
            self._ignored_headers_idx[self.current_index] = []
            for i in range(self._ncols):
                cell_value = self.current_sheet.get_cell_value(i+1, 1)
                if cell_value:
                    self._headers[self.current_index].append(self.current_sheet.get_cell_value(i+1, 1)[1])
                else:
                    logging.warn("Empty header in %s" % self.file_path)
                    self._headers[self.current_index].append("--empty--")
        return self._headers[self.current_index]

    def __next__(self):
        """ Returns an OrderedDict : {'DESCRIPTOR': value, ...} """
        while self.skip_lines and self._row_index in self.skip_lines:
            self._row_index += 1
        if self._row_index >= self._nrows:
            raise StopIteration
        row_dict = OrderedDict()
        for i in range(self._ncols):
            if i in self._ignored_headers_idx[self.current_index]:
                continue
            cell_value = self.current_sheet.get_cell_value(i+1, self._row_index+1)
            if cell_value and cell_value[0] == 'formula' and cell_value[1]:
                raise ValueError(_("The ODS file contains formula. Please convert them to raw values before importing the file."))
            row_dict[self.get_headers()[i]] = cell_value and cell_value[1] or ""
        self._row_index += 1
        return row_dict

    def current_sheet_name(self):
        return "??" # FIXME: self.current_sheet.title
